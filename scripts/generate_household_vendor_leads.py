#!/usr/bin/env python3
"""Generate filterable Excel vendor lead sheet for ScanV Household services card."""

import json
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CAPTURED = date.today().isoformat()
OUT = "/Users/samir/Downloads/scanverse/data/vendor-research/ScanV-Household-Vendors-Pune-PCMC.xlsx"
JSON_OUT = "/Users/samir/Downloads/scanverse/supabase/functions/_shared/vendor-leads-data.json"

SERVICES = [
    ("household", "Deep cleaning", "pink", "hh-bathroom-deep", "Bathroom Deep Clean"),
    ("household", "Deep cleaning", "pink", "hh-kitchen-deep", "Kitchen Deep Clean"),
    ("household", "Deep cleaning", "pink", "hh-flat-clean", "Full Flat Cleaning"),
    ("household", "Deep cleaning", "pink", "hh-care-plan", "Bathroom Care Plan"),
    ("household", "Deep cleaning", "pink", "hh-quick-clean", "Quick Clean"),
    ("household", "Deep cleaning", "pink", "hh-sofa-clean", "Sofa & Upholstery Clean"),
    ("household", "Home help", "green", "hh-house-help", "House Help"),
    ("household", "Home help", "green", "hh-dishwashing", "Dishwashing"),
    ("household", "Home help", "green", "hh-kitchen-help", "Kitchen Tidy-Up"),
    ("household", "Home help", "green", "hh-fan-clean", "Fan Cleaning"),
    ("household", "Home help", "green", "hh-window-clean", "Window Cleaning"),
    ("household", "Home help", "green", "hh-laundry", "Laundry Help"),
    ("household", "Home help", "green", "hh-bathroom-help", "Bathroom Refresh"),
    ("household", "Home help", "green", "hh-ironing", "Ironing & Pressing"),
]

# vendor dict + list of ScanV service IDs they can fulfill
VENDORS = [
    {
        "business_name": "Shreyash Deep Cleaning",
        "contact_person": "Ravindra Fodse",
        "shop_office": "Shreyash Deep Cleaning",
        "building": "Dudhal Building",
        "street": "Near Sonai Mangal Karayalaya, 16 No. Bus Stop, Kaveri Nagar, Kalewadi Phata",
        "area": "Wakad",
        "city": "Pune",
        "pin": "411057",
        "state": "Maharashtra",
        "phone1": "+91 8805839885",
        "phone2": "+91 9637913676",
        "email1": "service@shreyashdeepcleaning.com",
        "email2": "shreyashdeepcleanig@gmail.com",
        "website": "https://www.shreyashdeepcleaning.com/",
        "maps_name": "Shreyash Deep Cleaning Wakad",
        "services_offered": "Home deep clean, bathroom, kitchen, sofa, office cleaning",
        "rating": "",
        "hours": "Mon-Sun 8:00 AM - 8:00 PM",
        "service_areas": "Wakad, Kalewadi, PCMC",
        "source": "Website contact page",
        "notes": "Owner listed on website",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-quick-clean", "hh-sofa-clean"],
    },
    {
        "business_name": "Saaf Makers",
        "contact_person": "",
        "shop_office": "The Evoq",
        "building": "206, The Evoq",
        "street": "Kalakhadak Road",
        "area": "Wakad",
        "city": "Pune",
        "pin": "411057",
        "state": "Maharashtra",
        "phone1": "+91 8484888693",
        "phone2": "",
        "email1": "saafmakers29@gmail.com",
        "email2": "",
        "website": "https://saafmakers.com/",
        "maps_name": "Saaf Makers Wakad",
        "services_offered": "Deep home cleaning, maintenance, commercial (McDonald's Wakad reference)",
        "rating": "Top-rated (website claim)",
        "hours": "",
        "service_areas": "Wakad, Kothrud, Baner, Hinjewadi, Magarpatta, Hadapsar, PCMC",
        "source": "Website",
        "notes": "",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-quick-clean"],
    },
    {
        "business_name": "Deep Cleaning Pune",
        "contact_person": "",
        "shop_office": "Hermes Vishal Co-op Housing Society",
        "building": "B-4, Hermes Vishal",
        "street": "Lane No. 7, North Main Road",
        "area": "Koregaon Park",
        "city": "Pune",
        "pin": "411001",
        "state": "Maharashtra",
        "phone1": "+91 9975708557",
        "phone2": "",
        "email1": "info@deepcleanpune.com",
        "email2": "",
        "website": "https://deepcleanpune.com/",
        "maps_name": "Deep Cleaning Pune - Home Deep Cleaning in Pimpri Chinchwad, Wakad",
        "services_offered": "Home/office/commercial deep clean, sofa, carpet, kitchen, bathroom, PCMC",
        "rating": "Google Maps listing (Wakad search)",
        "hours": "24x7 contact",
        "service_areas": "Pune, PCMC, Wakad, Hinjewadi, Baner, Kharadi, Aundh, Kothrud",
        "source": "Website + Google Maps",
        "notes": "HQ Koregaon Park; serves Wakad/PCMC on-site",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-quick-clean", "hh-care-plan"],
    },
    {
        "business_name": "Dirt Blaster Cleaning Services",
        "contact_person": "",
        "shop_office": "Shri Nagari Society",
        "building": "Shop No D-6, Shri Nagari Society",
        "street": "Near Tupe Corner, Behind Amanora Mall, Malwadi Road",
        "area": "Hadapsar",
        "city": "Pune",
        "pin": "411028",
        "state": "Maharashtra",
        "phone1": "+91 7350321321",
        "phone2": "+91 8888086179",
        "email1": "support@dirtblastercleaningservices.com",
        "email2": "customercare@dirtbalster.in",
        "website": "https://www.dirtblastercleaningservices.com/",
        "maps_name": "Dirt Blaster Cleaning Services Pune",
        "services_offered": "Home deep clean, sofa, kitchen, bathroom, carpet, mattress, window, pest control",
        "rating": "4.9 (website claim)",
        "hours": "",
        "service_areas": "Pune, Pimpri Chinchwad, Wakad",
        "source": "Website contact",
        "notes": "Founded 2015",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-window-clean", "hh-quick-clean"],
    },
    {
        "business_name": "AS Deep Cleaning Services",
        "contact_person": "Rukhsana",
        "shop_office": "AS Deep Cleaning Services",
        "building": "",
        "street": "",
        "area": "Pimpri Chinchwad",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "+91 8087100195",
        "phone2": "",
        "email1": "rukhsana061991@gmail.com",
        "email2": "",
        "website": "https://www.deepcleaningservices.co.in/",
        "maps_name": "AS Deep Cleaning Services PCMC",
        "services_offered": "Home deep clean, office, carpet, bathroom, sofa, housekeeping, floor cleaning",
        "rating": "",
        "hours": "Mon-Sun 8:00 AM - 8:00 PM",
        "service_areas": "Pimpri Chinchwad, Pune",
        "source": "Website",
        "notes": "Street address not published — confirm on call",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-house-help"],
    },
    {
        "business_name": "B B Tule Housekeeping And Deep Cleaning Services",
        "contact_person": "",
        "shop_office": "World of Mother, Jai Ganesh Vision",
        "building": "E-111, World of Mother",
        "street": "Jai Ganesh Vision",
        "area": "Akurdi",
        "city": "Pimpri Chinchwad",
        "pin": "411035",
        "state": "Maharashtra",
        "phone1": "+91 9325106205",
        "phone2": "+91 9325106205",
        "email1": "",
        "email2": "",
        "website": "https://wakad.in/item/b-b-tule-housekeeping-services/",
        "maps_name": "B B Tule Housekeeping Services",
        "services_offered": "Housekeeping, deep clean, sofa, carpet, bathroom, kitchen, glass, office",
        "rating": "5.0/5 (Sulekha, 11 reviews)",
        "hours": "Mon-Sat 9:00 AM - 6:00 PM",
        "service_areas": "Akurdi, Pune, PCMC, Wakad",
        "source": "Sulekha + Wakad.in + LinkedIn",
        "notes": "Also listed E-111 A, Akurdi",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-window-clean", "hh-house-help"],
    },
    {
        "business_name": "Malhar Housekeeping",
        "contact_person": "",
        "shop_office": "Malhar Housekeeping",
        "building": "Near Pooja Hospital",
        "street": "Near Bank of Maharashtra, Gangarde Nagar",
        "area": "Pimple Gurav",
        "city": "Pimpri Chinchwad",
        "pin": "411061",
        "state": "Maharashtra",
        "phone1": "+91 8668639996",
        "phone2": "",
        "email1": "malharhousekeeping@gmail.com",
        "email2": "",
        "website": "https://malharhousekeeping.com/",
        "maps_name": "Malhar Housekeeping Pimple Gurav",
        "services_offered": "Housekeeping products & cleaning services (Pune, Pimple Gurav, Sangvi)",
        "rating": "",
        "hours": "",
        "service_areas": "Pimple Gurav, Sangvi, Pune",
        "source": "Website contact",
        "notes": "Also sells cleaning products",
        "service_ids": ["hh-house-help", "hh-kitchen-help", "hh-bathroom-help"],
    },
    {
        "business_name": "Talatkar Cleaning Company (Tulip Housekeeping)",
        "contact_person": "Omkar Talatkar",
        "shop_office": "Talatkar Cleaning Company",
        "building": "",
        "street": "",
        "area": "Pune & Pimpri Chinchwad",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "+91 8379041683",
        "phone2": "",
        "email1": "omkar.talatkar@gmail.com",
        "email2": "",
        "website": "https://talatkarcleaningcompany.com/",
        "maps_name": "Talatkar / Tulip Cleaning Pune",
        "services_offered": "Home/office deep clean, sofa, pest control, residential housekeeping",
        "rating": "",
        "hours": "",
        "service_areas": "Pune, Pimpri Chinchwad",
        "source": "Website",
        "notes": "Office address not on homepage — confirm on booking",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-window-clean", "hh-house-help"],
    },
    {
        "business_name": "S K W Enterprises",
        "contact_person": "Kumar",
        "shop_office": "S K W Enterprises",
        "building": "Sr No 48/4",
        "street": "Lane No. 12, Ganesh Nagar, Old Mundhwa Road",
        "area": "Kharadi",
        "city": "Pune",
        "pin": "411014",
        "state": "Maharashtra",
        "phone1": "+91 8796477425",
        "phone2": "+91 8329161666",
        "email1": "kumarpune91@gmail.com",
        "email2": "",
        "website": "https://skwenterprises.in/",
        "maps_name": "S K W Enterprises Kharadi",
        "services_offered": "Housekeeping, home deep clean, sofa shampoo, carpet, glass, sanitization",
        "rating": "",
        "hours": "",
        "service_areas": "Kharadi, Pune, PCMC",
        "source": "Website contact",
        "notes": "",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-window-clean"],
    },
    {
        "business_name": "DailyClean India Group",
        "contact_person": "",
        "shop_office": "Arambha Office",
        "building": "Office No. 5, 1st Floor, Arambha",
        "street": "Datta Mandir Road",
        "area": "Wakad",
        "city": "Pune",
        "pin": "411057",
        "state": "Maharashtra",
        "phone1": "+91 7709902201",
        "phone2": "+91 8460308711",
        "email1": "",
        "email2": "",
        "website": "https://dailyclean.co.in/",
        "maps_name": "DailyClean - Cleaning services in Pune",
        "services_offered": "Home deep clean, mattress, bathroom, commercial cleaning",
        "rating": "Google Maps Wakad search",
        "hours": "Mon-Sat 10:00 AM - 7:00 PM",
        "service_areas": "Wakad, Pune, PCMC, Hadapsar",
        "source": "Website + Google Maps + bdir.in",
        "notes": "Email not public — use phone/WhatsApp",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-quick-clean"],
    },
    {
        "business_name": "A3 Care Deep Cleaning & Housekeeping Services",
        "contact_person": "",
        "shop_office": "A3 Care (service-area business)",
        "building": "",
        "street": "",
        "area": "Wakad / PCMC",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "+91 7822954195",
        "phone2": "+91 8605166808",
        "email1": "a3caredeepcleaning@gmail.com",
        "email2": "",
        "website": "https://a3caredeepcleaningservices.com/",
        "maps_name": "A3 Care - Home | Office Cleaning Services Wakad PCMC",
        "services_offered": "Home/flat deep clean, kitchen, toilet, sofa/carpet, housekeeping, marble polish",
        "rating": "Google Maps Wakad search",
        "hours": "",
        "service_areas": "Wakad, Baner, Balewadi, Hinjewadi, Ravet, Akurdi, Nigdi, PCMC",
        "source": "Website + Google Maps",
        "notes": "No fixed shop address published",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-care-plan", "hh-house-help"],
    },
    {
        "business_name": "Hello Help (Tidy Troop Private Limited)",
        "contact_person": "",
        "shop_office": "Hello Help (app platform)",
        "building": "",
        "street": "",
        "area": "Pimpri Chinchwad",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "",
        "phone2": "",
        "email1": "service@hellohelp.in",
        "email2": "",
        "website": "https://hellohelp.in/",
        "maps_name": "Hello Help maid service PCMC",
        "services_offered": "On-demand verified maids: broom, mop, dishes, bedding, laundry, dusting, veg chop, 1 bathroom",
        "rating": "",
        "hours": "App booking",
        "service_areas": "Hinjewadi, Wakad, Chinchwad, Akurdi, PCMC",
        "source": "Website",
        "notes": "Platform vendor — onboard as partner not shop",
        "service_ids": ["hh-house-help", "hh-dishwashing", "hh-kitchen-help", "hh-laundry", "hh-bathroom-help"],
    },
    {
        "business_name": "Maid Services Pune",
        "contact_person": "Jayraj Vastushilp office",
        "shop_office": "Jayraj Vastushilp Housing Society",
        "building": "Jayraj Vastushilp Housing Society",
        "street": "Wadgaon Bridge, Opposite Abhiruchi Mall",
        "area": "Wadgaon Budruk",
        "city": "Pune",
        "pin": "411041",
        "state": "Maharashtra",
        "phone1": "+91 6376634479",
        "phone2": "",
        "email1": "maidserviceinpune@gmail.com",
        "email2": "",
        "website": "https://maidservicespune.com/",
        "maps_name": "Maid Services Pune Wakad",
        "services_offered": "Full-time/part-time/live-in maids, cooking, cleaning, childcare, elderly care",
        "rating": "",
        "hours": "Sun-Sat 8:00 AM - 5:00 PM",
        "service_areas": "Wakad, Hinjewadi, Baner, Balewadi, PCMC, Kharadi, Hadapsar",
        "source": "Website",
        "notes": "Agency model — placement not hourly app",
        "service_ids": ["hh-house-help", "hh-dishwashing", "hh-kitchen-help", "hh-laundry", "hh-ironing", "hh-bathroom-help", "hh-fan-clean"],
    },
    {
        "business_name": "Beyond Home Services",
        "contact_person": "",
        "shop_office": "Fortune Business Center",
        "building": "505, Fortune Business Center",
        "street": "Vishnudev Nagar",
        "area": "Wakad",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "+91 8459507288",
        "phone2": "",
        "email1": "",
        "email2": "",
        "website": "https://beyondhome.services/",
        "maps_name": "Beyond Home Services Wakad",
        "services_offered": "Maids, cooks, nannies, cleaning, patient care — Western Pune",
        "rating": "",
        "hours": "Mon-Sat 9:30 AM - 6:30 PM",
        "service_areas": "Wakad, Baner, Balewadi, Pimple Saudagar, Hinjewadi, Pimple Gurav, Tathawade",
        "source": "Website",
        "notes": "PIN not listed — Wakad 411057 area typical",
        "service_ids": ["hh-house-help", "hh-dishwashing", "hh-kitchen-help", "hh-laundry", "hh-ironing", "hh-bathroom-help"],
    },
    {
        "business_name": "Snabbit",
        "contact_person": "",
        "shop_office": "Snabbit (app platform)",
        "building": "",
        "street": "",
        "area": "Wakad / Hinjewadi",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "",
        "phone2": "",
        "email1": "",
        "email2": "",
        "website": "https://www.snabbit.com/house-help/pune/wakad",
        "maps_name": "Snabbit house help Wakad",
        "services_offered": "Hourly verified female helpers: deep clean, mop, dust, dishes, laundry, kitchen, bathroom",
        "rating": "",
        "hours": "7:00 AM - 7:45 PM",
        "service_areas": "Wakad, Hinjewadi, Baner, Kharadi, Hadapsar, Viman Nagar",
        "source": "Website",
        "notes": "From ₹99/hr app pricing — competitor/partner reference",
        "service_ids": ["hh-house-help", "hh-dishwashing", "hh-kitchen-help", "hh-laundry", "hh-bathroom-help", "hh-fan-clean", "hh-window-clean"],
    },
    {
        "business_name": "Deep Cleaning Services Pune (Housekeeping division)",
        "contact_person": "",
        "shop_office": "",
        "building": "",
        "street": "",
        "area": "Pimpri Chinchwad",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "+91 9960228504",
        "phone2": "",
        "email1": "info@deepcleaningservicespune.com",
        "email2": "",
        "website": "https://deepcleaningpune.com/",
        "maps_name": "Deep cleaning housekeeping PCMC",
        "services_offered": "Housekeeping, toilet/bathroom deep clean, kitchen deep clean, sofa/carpet, office commercial",
        "rating": "",
        "hours": "",
        "service_areas": "PCMC, Wakad, Hinjewadi, Baner, Balewadi, Pimple Saudagar, Rahatani, Kalewadi",
        "source": "Website",
        "notes": "Different entity from deepcleanpune.com — verify legal name on call",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean", "hh-care-plan", "hh-house-help"],
    },
    {
        "business_name": "Varsha Deep Cleaning Service Wakad",
        "contact_person": "",
        "shop_office": "",
        "building": "",
        "street": "",
        "area": "Wakad",
        "city": "Pune",
        "pin": "411057",
        "state": "Maharashtra",
        "phone1": "",
        "phone2": "",
        "email1": "",
        "email2": "",
        "website": "",
        "maps_name": "Varsha deep cleaning service.wakad,pune",
        "services_offered": "Deep cleaning (Google Maps listing)",
        "rating": "Google Maps",
        "hours": "",
        "service_areas": "Wakad",
        "source": "Google Maps search: deep cleaning Wakad Pune",
        "notes": "Phone/address — open listing in Maps and verify",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-quick-clean"],
    },
    {
        "business_name": "Punekar Deep Cleaning",
        "contact_person": "",
        "shop_office": "",
        "building": "",
        "street": "",
        "area": "Wakad / Pune",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "",
        "phone2": "",
        "email1": "",
        "email2": "",
        "website": "",
        "maps_name": "Punekar Deep Cleaning",
        "services_offered": "Deep cleaning (Google Maps + website link in Maps)",
        "rating": "Google Maps",
        "hours": "",
        "service_areas": "Pune",
        "source": "Google Maps",
        "notes": "Verify contact from Maps listing",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-sofa-clean"],
    },
    {
        "business_name": "Sai Facility And Deep Cleaning Service Wakad",
        "contact_person": "",
        "shop_office": "",
        "building": "",
        "street": "",
        "area": "Wakad",
        "city": "Pune",
        "pin": "411057",
        "state": "Maharashtra",
        "phone1": "",
        "phone2": "",
        "email1": "",
        "email2": "",
        "website": "",
        "maps_name": "Sai facility and Deep cleaning service wakad",
        "services_offered": "Facility & deep cleaning",
        "rating": "Google Maps",
        "hours": "",
        "service_areas": "Wakad",
        "source": "Google Maps",
        "notes": "Has website link in Maps — verify on listing",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean"],
    },
    {
        "business_name": "Prime Clean Deep Cleaning & Housekeeping Services",
        "contact_person": "",
        "shop_office": "",
        "building": "",
        "street": "",
        "area": "Wakad / PCMC",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "",
        "phone2": "",
        "email1": "",
        "email2": "",
        "website": "",
        "maps_name": "Prime Clean Deep Cleaning & Housekeeping Services",
        "services_offered": "Deep cleaning & housekeeping",
        "rating": "Google Maps",
        "hours": "",
        "service_areas": "PCMC, Pune",
        "source": "Google Maps",
        "notes": "Verify phone/email from Maps listing",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-house-help"],
    },
    {
        "business_name": "TS Deep House Cleaning Wala",
        "contact_person": "",
        "shop_office": "",
        "building": "",
        "street": "",
        "area": "Wakad",
        "city": "Pune",
        "pin": "",
        "state": "Maharashtra",
        "phone1": "",
        "phone2": "",
        "email1": "",
        "email2": "",
        "website": "",
        "maps_name": "TS Deep House Cleaning Wala",
        "services_offered": "Home deep cleaning",
        "rating": "Google Maps",
        "hours": "",
        "service_areas": "Wakad, Pune",
        "source": "Google Maps",
        "notes": "Verify contact from Maps listing",
        "service_ids": ["hh-bathroom-deep", "hh-kitchen-deep", "hh-flat-clean", "hh-quick-clean"],
    },
]

HEADERS = [
    "ScanV Parent Card",
    "ScanV Sub-card",
    "ScanV Theme",
    "ScanV Service ID",
    "ScanV Service Name",
    "Business Name",
    "Contact Person",
    "Shop / Office Name",
    "Building / Society",
    "Street / Road",
    "Area / Locality",
    "City",
    "PIN Code",
    "State",
    "Phone Primary",
    "Phone Secondary",
    "Email Primary",
    "Email Secondary",
    "Website",
    "Google Maps Listing",
    "Vendor Services Offered",
    "Google Rating / Reviews",
    "Business Hours",
    "Service Areas Covered",
    "Lead Source",
    "Verification Notes",
    "Match Confidence",
    "Captured Date",
]

SERVICE_LOOKUP = {sid: row for row in SERVICES for sid in [row[3]]}


def full_address(v):
    parts = [v.get("building"), v.get("street"), v.get("area"), v.get("city"), v.get("pin"), v.get("state")]
    return ", ".join(p for p in parts if p)


def slugify(name):
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return s[:80] or "vendor"


PRICES = {
    "hh-bathroom-deep": 499, "hh-kitchen-deep": 599, "hh-flat-clean": 1999, "hh-care-plan": 1499,
    "hh-quick-clean": 149, "hh-sofa-clean": 249, "hh-house-help": 182, "hh-dishwashing": 99,
    "hh-kitchen-help": 149, "hh-fan-clean": 149, "hh-window-clean": 199, "hh-laundry": 149,
    "hh-bathroom-help": 199, "hh-ironing": 149,
}


def build_json_catalog():
    services = []
    for parent, sub, theme, sid, name in SERVICES:
        services.append({
            "id": sid,
            "parent_card_id": parent,
            "parent_card_label": "Household services",
            "sub_card": sub,
            "theme": theme,
            "name": name,
            "price_inr": PRICES.get(sid),
        })

    vendors = []
    for v in VENDORS:
        confidence = "high" if v.get("phone1") or v.get("email1") else "verify_maps"
        vendors.append({
            "id": slugify(v["business_name"]),
            "business_name": v["business_name"],
            "contact_person": v.get("contact_person") or "",
            "shop_office": v.get("shop_office") or "",
            "address": {
                "building": v.get("building") or "",
                "street": v.get("street") or "",
                "area": v.get("area") or "",
                "city": v.get("city") or "",
                "pin": v.get("pin") or "",
                "state": v.get("state") or "Maharashtra",
                "full": full_address(v),
            },
            "phones": [p for p in [v.get("phone1"), v.get("phone2")] if p],
            "emails": [e for e in [v.get("email1"), v.get("email2")] if e],
            "website": v.get("website") or "",
            "maps_name": v.get("maps_name") or "",
            "services_offered": v.get("services_offered") or "",
            "rating": v.get("rating") or "",
            "hours": v.get("hours") or "",
            "service_areas": v.get("service_areas") or "",
            "source": v.get("source") or "",
            "notes": v.get("notes") or "",
            "confidence": confidence,
            "service_ids": list(v.get("service_ids") or []),
        })

    return {
        "meta": {
            "captured_at": CAPTURED,
            "market": "Pune / PCMC",
            "version": 1,
            "vendor_count": len(vendors),
            "service_count": len(services),
        },
        "cards": [{
            "id": "household",
            "label": "Household services",
            "icon": "🧹",
            "sub_cards": [
                {"id": "deep-cleaning", "label": "Deep cleaning", "theme": "pink"},
                {"id": "home-help", "label": "Home help", "theme": "green"},
            ],
        }],
        "services": services,
        "vendors": vendors,
    }


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    widths = [14, 14, 10, 16, 22, 28, 14, 22, 22, 24, 16, 12, 10, 14, 16, 16, 24, 24, 28, 28, 32, 14, 16, 24, 18, 28, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = Workbook()

    # Reference sheet
    ws_ref = wb.active
    ws_ref.title = "ScanV Services"
    ref_headers = ["Parent Card", "Sub-card", "Theme", "Service ID", "Service Name", "ScanV Price (₹)"]
    for c, h in enumerate(ref_headers, 1):
        cell = ws_ref.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.font = Font(color="FFFFFF", bold=True)
    for r, (parent, sub, theme, sid, name) in enumerate(SERVICES, 2):
        ws_ref.append(["Household services", sub, theme, sid, name, PRICES.get(sid, "")])
    ws_ref.freeze_panes = "A2"
    ws_ref.auto_filter.ref = f"A1:F{len(SERVICES)+1}"

    # Vendor leads — one row per vendor × service
    ws = wb.create_sheet("Vendor Leads")
    style_sheet(ws)
    row = 2
    for v in VENDORS:
        confidence = "High" if v["phone1"] or v["email1"] else "Verify in Maps"
        for sid in v["service_ids"]:
            parent, sub, theme, _, sname = SERVICE_LOOKUP[sid]
            ws.append([
                "Household services",
                sub,
                theme,
                sid,
                sname,
                v["business_name"],
                v["contact_person"],
                v["shop_office"],
                v["building"],
                v["street"],
                v["area"],
                v["city"],
                v["pin"],
                v["state"],
                v["phone1"],
                v["phone2"],
                v["email1"],
                v["email2"],
                v["website"],
                v["maps_name"],
                v["services_offered"],
                v["rating"],
                v["hours"],
                v["service_areas"],
                v["source"],
                v["notes"],
                confidence,
                CAPTURED,
            ])
            row += 1

    # Vendor master (deduped)
    ws2 = wb.create_sheet("Vendor Master")
    master_headers = [
        "Business Name", "Contact Person", "Full Address", "City", "PIN", "State",
        "Phone Primary", "Phone Secondary", "Email Primary", "Email Secondary", "Website",
        "Google Maps Listing", "Services Offered", "Service Areas", "Source", "Notes", "Captured Date",
    ]
    for c, h in enumerate(master_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="5D4037")
        cell.font = Font(color="FFFFFF", bold=True)
    for v in VENDORS:
        ws2.append([
            v["business_name"], v["contact_person"], full_address(v), v["city"], v["pin"], v["state"],
            v["phone1"], v["phone2"], v["email1"], v["email2"], v["website"], v["maps_name"],
            v["services_offered"], v["service_areas"], v["source"], v["notes"], CAPTURED,
        ])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(master_headers))}{len(VENDORS)+1}"

    wb.save(OUT)

    catalog = build_json_catalog()
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"Wrote {OUT}")
    print(f"Wrote {JSON_OUT}")
    print(f"Vendors: {len(VENDORS)} | Service mapping rows: {row-2}")


if __name__ == "__main__":
    main()
