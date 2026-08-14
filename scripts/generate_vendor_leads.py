#!/usr/bin/env python3
"""Generate filterable Excel + JSON vendor lead catalog for all ScanV cards (Pune / PCMC)."""

import json
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vendor_leads_extra_vendors import EXTRA_VENDORS

ROOT = Path(__file__).resolve().parents[1]
CATALOG_JSON = ROOT / "scripts" / "scanv_catalog_extract.json"
OUT_XLSX = ROOT / "data" / "vendor-research" / "ScanV-All-Cards-Vendors-Pune-PCMC.xlsx"
JSON_OUT = ROOT / "supabase" / "functions" / "_shared" / "vendor-leads-data.json"
HOUSEHOLD_SCRIPT = ROOT / "scripts" / "generate_household_vendor_leads.py"

CAPTURED = date.today().isoformat()

HEADERS = [
    "ScanV Parent Card",
    "ScanV Sub-card",
    "ScanV Theme",
    "ScanV Service ID",
    "ScanV Service Name",
    "Business Name",
    "Contact Person",
    "Shop / Office Name",
    "Building / Society",
    "Street / Road",
    "Area / Locality",
    "City",
    "PIN Code",
    "State",
    "Phone Primary",
    "Phone Secondary",
    "Email Primary",
    "Email Secondary",
    "Website",
    "Google Maps Listing",
    "Vendor Services Offered",
    "Google Rating / Reviews",
    "Business Hours",
    "Service Areas Covered",
    "Lead Source",
    "Verification Notes",
    "Match Confidence",
    "Captured Date",
]


def slugify(name):
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return s[:80] or "vendor"


def full_address(v):
    parts = [v.get("building"), v.get("street"), v.get("area"), v.get("city"), v.get("pin"), v.get("state")]
    return ", ".join(p for p in parts if p)


def load_household_vendors():
    ns = {}
    with open(HOUSEHOLD_SCRIPT, encoding="utf-8") as f:
        code = f.read()
    exec(compile(code, str(HOUSEHOLD_SCRIPT), "exec"), ns)  # noqa: S102
    return ns.get("VENDORS", [])


def vendor_row_to_json(v, used_ids):
    base = slugify(v["business_name"])
    vid = base
    n = 2
    while vid in used_ids:
        vid = f"{base}-{n}"
        n += 1
    used_ids.add(vid)
    confidence = "high" if v.get("phone1") or v.get("email1") else "verify_maps"
    return {
        "id": vid,
        "business_name": v["business_name"],
        "contact_person": v.get("contact_person") or "",
        "shop_office": v.get("shop_office") or "",
        "address": {
            "building": v.get("building") or "",
            "street": v.get("street") or "",
            "area": v.get("area") or "",
            "city": v.get("city") or "",
            "pin": v.get("pin") or "",
            "state": v.get("state") or "Maharashtra",
            "full": full_address(v),
        },
        "phones": [p for p in [v.get("phone1"), v.get("phone2")] if p],
        "emails": [e for e in [v.get("email1"), v.get("email2")] if e],
        "website": v.get("website") or "",
        "maps_name": v.get("maps_name") or "",
        "services_offered": v.get("services_offered") or "",
        "rating": v.get("rating") or "",
        "hours": v.get("hours") or "",
        "service_areas": v.get("service_areas") or "",
        "source": v.get("source") or "",
        "notes": v.get("notes") or "",
        "confidence": confidence,
        "service_ids": list(v.get("service_ids") or []),
    }


def build_catalog():
    with open(CATALOG_JSON, encoding="utf-8") as f:
        extracted = json.load(f)
    cards = extracted["cards"]
    services = extracted["services"]
    service_lookup = {s["id"]: s for s in services}

    used_ids = set()
    raw_vendors = load_household_vendors() + EXTRA_VENDORS
    vendors = []
    for v in raw_vendors:
        row = vendor_row_to_json(v, used_ids)
        row["service_ids"] = [sid for sid in row["service_ids"] if sid in service_lookup]
        if row["service_ids"]:
            vendors.append(row)

    return {
        "meta": {
            "captured_at": CAPTURED,
            "market": "Pune / PCMC",
            "version": 2,
            "vendor_count": len(vendors),
            "service_count": len(services),
            "card_count": len(cards),
        },
        "cards": cards,
        "services": services,
        "vendors": vendors,
    }


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    widths = [14, 14, 10, 16, 22, 28, 14, 22, 22, 24, 16, 12, 10, 14, 16, 16, 24, 24, 28, 28, 32, 14, 16, 24, 18, 28, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    catalog = build_catalog()
    service_lookup = {s["id"]: s for s in catalog["services"]}

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_ref = wb.active
    ws_ref.title = "ScanV Services"
    ref_headers = ["Parent Card", "Sub-card", "Theme", "Service ID", "Service Name", "ScanV Price (₹)"]
    for c, h in enumerate(ref_headers, 1):
        cell = ws_ref.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.font = Font(color="FFFFFF", bold=True)
    for s in catalog["services"]:
        ws_ref.append([
            s["parent_card_label"], s["sub_card"], s["theme"], s["id"], s["name"], s.get("price_inr", ""),
        ])
    ws_ref.freeze_panes = "A2"
    ws_ref.auto_filter.ref = f"A1:F{len(catalog['services']) + 1}"

    ws = wb.create_sheet("Vendor Leads")
    style_sheet(ws)
    row = 2
    for v in catalog["vendors"]:
        confidence = "High" if v["phones"] or v["emails"] else "Verify in Maps"
        for sid in v["service_ids"]:
            svc = service_lookup[sid]
            ws.append([
                svc["parent_card_label"],
                svc["sub_card"],
                svc["theme"],
                sid,
                svc["name"],
                v["business_name"],
                v["contact_person"],
                v["shop_office"],
                v["address"]["building"],
                v["address"]["street"],
                v["address"]["area"],
                v["address"]["city"],
                v["address"]["pin"],
                v["address"]["state"],
                v["phones"][0] if v["phones"] else "",
                v["phones"][1] if len(v["phones"]) > 1 else "",
                v["emails"][0] if v["emails"] else "",
                v["emails"][1] if len(v["emails"]) > 1 else "",
                v["website"],
                v["maps_name"],
                v["services_offered"],
                v["rating"],
                v["hours"],
                v["service_areas"],
                v["source"],
                v["notes"],
                confidence,
                CAPTURED,
            ])
            row += 1

    ws2 = wb.create_sheet("Vendor Master")
    master_headers = [
        "Lead ID", "Business Name", "Contact Person", "Full Address", "City", "PIN", "State",
        "Phone Primary", "Email Primary", "Website", "Google Maps Listing", "Services Offered",
        "Service Areas", "ScanV Service IDs", "Source", "Notes", "Confidence", "Captured Date",
    ]
    for c, h in enumerate(master_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="5D4037")
        cell.font = Font(color="FFFFFF", bold=True)
    for v in catalog["vendors"]:
        ws2.append([
            v["id"], v["business_name"], v["contact_person"], v["address"]["full"],
            v["address"]["city"], v["address"]["pin"], v["address"]["state"],
            v["phones"][0] if v["phones"] else "",
            v["emails"][0] if v["emails"] else "",
            v["website"], v["maps_name"], v["services_offered"], v["service_areas"],
            ", ".join(v["service_ids"]), v["source"], v["notes"], v["confidence"], CAPTURED,
        ])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(master_headers))}{len(catalog['vendors']) + 1}"

    wb.save(OUT_XLSX)

    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {JSON_OUT}")
    print(
        f"Cards: {catalog['meta']['card_count']} | Services: {catalog['meta']['service_count']} | "
        f"Vendors: {catalog['meta']['vendor_count']} | Mapping rows: {row - 2}"
    )


if __name__ == "__main__":
    main()
